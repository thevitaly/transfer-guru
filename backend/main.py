from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
import pandas as pd
from openpyxl import load_workbook
import os
from typing import Optional
import tempfile
import shutil

app = FastAPI(title="Transfer Guru API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Global dataframe storage
DATA_STORE = {
    "df": None,
    "filename": None
}

# Column mappings
COLUMNS = {
    "legal_name": 0,
    "brand_name": 1,
    "acquirer": 2,
    "currency": 5,
    "amount": 6,
    "fee": 8,
    "psp_buy_fee": 11,
    "type": 14,
    "status": 15,
}

HEADER_ROW = 16  # 0-indexed: row 16 is headers, data starts at 17


def parse_european_number(value):
    """Convert European format numbers (comma as decimal) to float"""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        # Skip formula cells
        if value.startswith("="):
            return 0.0
        # Replace comma with dot for European format
        value = value.replace(",", ".")
        try:
            return float(value)
        except ValueError:
            return 0.0
    return 0.0


def load_xlsx_data(file_path: str) -> pd.DataFrame:
    """Load xlsx file and return cleaned DataFrame"""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    data = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=HEADER_ROW + 2, values_only=True)):
        # Skip rows with formulas in Type column (summary rows at bottom)
        type_val = row[COLUMNS["type"]]
        status_val = row[COLUMNS["status"]]

        if type_val is None or status_val is None:
            continue
        if isinstance(type_val, str) and type_val.startswith("="):
            continue
        if not isinstance(type_val, str):
            continue

        # Normalize type and status
        type_lower = type_val.lower().strip()
        status_lower = status_val.lower().strip() if isinstance(status_val, str) else str(status_val).lower().strip()

        data.append({
            "legal_name": str(row[COLUMNS["legal_name"]] or ""),
            "brand_name": str(row[COLUMNS["brand_name"]] or ""),
            "acquirer": str(row[COLUMNS["acquirer"]] or ""),
            "currency": str(row[COLUMNS["currency"]] or ""),
            "amount": parse_european_number(row[COLUMNS["amount"]]),
            "fee": parse_european_number(row[COLUMNS["fee"]]),
            "psp_buy_fee": parse_european_number(row[COLUMNS["psp_buy_fee"]]),
            "type": type_lower,
            "status": status_lower,
        })

    wb.close()
    return pd.DataFrame(data)


def filter_by_operation_type(df: pd.DataFrame, op_type: int) -> pd.DataFrame:
    """
    Filter data by operation type:
    1 = purchase (paid, refunded, chargedback)
    2 = refund (success)
    3 = chargeback (success)
    4 = payout (success)
    """
    if op_type == 1:
        return df[
            (df["type"] == "purchase") &
            (df["status"].isin(["paid", "refunded", "chargedback"]))
        ]
    elif op_type == 2:
        return df[(df["type"] == "refund") & (df["status"] == "success")]
    elif op_type == 3:
        return df[(df["type"] == "chargeback") & (df["status"] == "success")]
    elif op_type == 4:
        return df[(df["type"] == "payout") & (df["status"] == "success")]
    else:
        raise ValueError(f"Unknown operation type: {op_type}")


def build_pivot_by_acquirer(df: pd.DataFrame) -> dict:
    """
    Build pivot table grouped by: Acquirer → Legal Name → Currency
    Type 1 format
    """
    if df.empty:
        return {"groups": [], "totals": {"amount": 0, "fee": 0, "psp_buy_fee": 0, "count": 0}}

    grouped = df.groupby(["acquirer", "legal_name", "currency"]).agg(
        amount=("amount", "sum"),
        fee=("fee", "sum"),
        psp_buy_fee=("psp_buy_fee", "sum"),
        count=("amount", "count")
    ).reset_index()

    result = {"groups": [], "totals": {
        "amount": round(grouped["amount"].sum(), 2),
        "fee": round(grouped["fee"].sum(), 2),
        "psp_buy_fee": round(grouped["psp_buy_fee"].sum(), 2),
        "count": int(grouped["count"].sum())
    }}

    # Group by acquirer
    for acquirer in grouped["acquirer"].unique():
        acq_data = grouped[grouped["acquirer"] == acquirer]
        acq_group = {
            "acquirer": acquirer,
            "merchants": [],
            "subtotals": {
                "amount": round(acq_data["amount"].sum(), 2),
                "fee": round(acq_data["fee"].sum(), 2),
                "psp_buy_fee": round(acq_data["psp_buy_fee"].sum(), 2),
                "count": int(acq_data["count"].sum())
            }
        }

        # Group by legal_name within acquirer
        for legal_name in acq_data["legal_name"].unique():
            merchant_data = acq_data[acq_data["legal_name"] == legal_name]
            merchant_group = {
                "legal_name": legal_name,
                "currencies": [],
                "subtotals": {
                    "amount": round(merchant_data["amount"].sum(), 2),
                    "fee": round(merchant_data["fee"].sum(), 2),
                    "psp_buy_fee": round(merchant_data["psp_buy_fee"].sum(), 2),
                    "count": int(merchant_data["count"].sum())
                }
            }

            for _, row in merchant_data.iterrows():
                merchant_group["currencies"].append({
                    "currency": row["currency"],
                    "amount": round(row["amount"], 2),
                    "fee": round(row["fee"], 2),
                    "psp_buy_fee": round(row["psp_buy_fee"], 2),
                    "count": int(row["count"])
                })

            acq_group["merchants"].append(merchant_group)

        result["groups"].append(acq_group)

    return result


def build_pivot_by_merchant(df: pd.DataFrame) -> dict:
    """
    Build pivot table grouped by: Legal Name → Acquirer → Currency
    Type 2 format
    """
    if df.empty:
        return {"groups": [], "totals": {"amount": 0, "fee": 0, "psp_buy_fee": 0, "count": 0}}

    grouped = df.groupby(["legal_name", "acquirer", "currency"]).agg(
        amount=("amount", "sum"),
        fee=("fee", "sum"),
        psp_buy_fee=("psp_buy_fee", "sum"),
        count=("amount", "count")
    ).reset_index()

    result = {"groups": [], "totals": {
        "amount": round(grouped["amount"].sum(), 2),
        "fee": round(grouped["fee"].sum(), 2),
        "psp_buy_fee": round(grouped["psp_buy_fee"].sum(), 2),
        "count": int(grouped["count"].sum())
    }}

    # Group by legal_name
    for legal_name in grouped["legal_name"].unique():
        merchant_data = grouped[grouped["legal_name"] == legal_name]
        merchant_group = {
            "legal_name": legal_name,
            "acquirers": [],
            "subtotals": {
                "amount": round(merchant_data["amount"].sum(), 2),
                "fee": round(merchant_data["fee"].sum(), 2),
                "psp_buy_fee": round(merchant_data["psp_buy_fee"].sum(), 2),
                "count": int(merchant_data["count"].sum())
            }
        }

        # Group by acquirer within merchant
        for acquirer in merchant_data["acquirer"].unique():
            acq_data = merchant_data[merchant_data["acquirer"] == acquirer]
            acq_group = {
                "acquirer": acquirer,
                "currencies": [],
                "subtotals": {
                    "amount": round(acq_data["amount"].sum(), 2),
                    "fee": round(acq_data["fee"].sum(), 2),
                    "psp_buy_fee": round(acq_data["psp_buy_fee"].sum(), 2),
                    "count": int(acq_data["count"].sum())
                }
            }

            for _, row in acq_data.iterrows():
                acq_group["currencies"].append({
                    "currency": row["currency"],
                    "amount": round(row["amount"], 2),
                    "fee": round(row["fee"], 2),
                    "psp_buy_fee": round(row["psp_buy_fee"], 2),
                    "count": int(row["count"])
                })

            merchant_group["acquirers"].append(acq_group)

        result["groups"].append(merchant_group)

    return result


@app.get("/")
async def root():
    return {"status": "ok", "message": "Transfer Guru API"}


@app.get("/api/status")
async def get_status():
    return {
        "loaded": DATA_STORE["df"] is not None,
        "filename": DATA_STORE["filename"],
        "row_count": len(DATA_STORE["df"]) if DATA_STORE["df"] is not None else 0
    }


@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    """Upload and parse xlsx file"""
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")

    # Save to temp file
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        shutil.copyfileobj(file.file, tmp)
        tmp_path = tmp.name

    try:
        df = load_xlsx_data(tmp_path)
        DATA_STORE["df"] = df
        DATA_STORE["filename"] = file.filename

        # Get stats
        type_counts = df.groupby(["type", "status"]).size().reset_index(name="count")
        stats = type_counts.to_dict(orient="records")

        return {
            "success": True,
            "filename": file.filename,
            "total_rows": len(df),
            "type_status_breakdown": stats
        }
    finally:
        os.unlink(tmp_path)


@app.post("/api/load-default")
async def load_default_file():
    """Load the default tab.xlsx file"""
    file_path = "/Users/lcc/transfer-guru/tab.xlsx"

    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Default file not found")

    df = load_xlsx_data(file_path)
    DATA_STORE["df"] = df
    DATA_STORE["filename"] = "tab.xlsx"

    # Get stats
    type_counts = df.groupby(["type", "status"]).size().reset_index(name="count")
    stats = type_counts.to_dict(orient="records")

    return {
        "success": True,
        "filename": "tab.xlsx",
        "total_rows": len(df),
        "type_status_breakdown": stats
    }


@app.get("/api/pivot")
async def get_pivot(
    operation_type: int,
    view_type: int = 1,  # 1 = by acquirer, 2 = by merchant
    currency: Optional[str] = None
):
    """
    Get pivot table data

    operation_type: 1-4 (purchase, refund, chargeback, payout)
    view_type: 1 = Acquirer→Legal Name→Currency, 2 = Legal Name→Acquirer→Currency
    currency: Optional filter by currency (e.g., "EUR")
    """
    if DATA_STORE["df"] is None:
        raise HTTPException(status_code=400, detail="No data loaded. Upload a file first.")

    df = DATA_STORE["df"].copy()

    # Filter by operation type
    df = filter_by_operation_type(df, operation_type)

    # Filter by currency if specified
    if currency:
        df = df[df["currency"] == currency.upper()]

    # Build pivot based on view type
    if view_type == 1:
        pivot = build_pivot_by_acquirer(df)
    else:
        pivot = build_pivot_by_merchant(df)

    return {
        "operation_type": operation_type,
        "view_type": view_type,
        "currency_filter": currency,
        "data": pivot
    }


@app.get("/api/currencies")
async def get_currencies():
    """Get list of unique currencies in the data"""
    if DATA_STORE["df"] is None:
        raise HTTPException(status_code=400, detail="No data loaded")

    currencies = sorted(DATA_STORE["df"]["currency"].unique().tolist())
    return {"currencies": currencies}


@app.get("/api/summary")
async def get_summary():
    """Get summary statistics for all operation types"""
    if DATA_STORE["df"] is None:
        raise HTTPException(status_code=400, detail="No data loaded")

    df = DATA_STORE["df"]

    summaries = []
    operation_names = {
        1: "Purchase (paid/refunded/chargedback)",
        2: "Refund (success)",
        3: "Chargeback (success)",
        4: "Payout (success)"
    }

    for op_type in [1, 2, 3, 4]:
        filtered = filter_by_operation_type(df, op_type)
        summaries.append({
            "operation_type": op_type,
            "name": operation_names[op_type],
            "count": len(filtered),
            "total_amount": round(filtered["amount"].sum(), 2),
            "total_fee": round(filtered["fee"].sum(), 2),
            "total_psp_buy_fee": round(filtered["psp_buy_fee"].sum(), 2)
        })

    return {"summaries": summaries}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8080)
